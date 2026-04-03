"""
generate_intercoder_materials.py
=================================
Generates intercoder reliability materials for the DarkSideofAI thesis.

Produces two files:
  output/intercoder/intercoder_coding_sheet.xlsx
      — spreadsheet for the second coder to fill in independently

  output/intercoder/domain_list_for_codebook.csv
      — domain list used to verify what the second coder receives

Run from project root:
    python3 src/generate_intercoder_materials.py
"""

import sqlite3
import csv
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                  Border, Side)
    from openpyxl.utils import get_column_letter
except ImportError:
    raise SystemExit("openpyxl not installed. Run: pip install openpyxl")

DB_PATH    = "data/scraping.db"
OUTPUT_DIR = Path("output/intercoder")

# URL-based splits (from 01_prepare_audience_overrides.py)
URL_SPLITS = {
    "clickworker.com": {
        "worker_path": "/clickworker/",
        "note": "Pages under /clickworker/ target annotators; all other pages target clients"
    },
    "prolific.com": {
        "worker_path": "/participants",
        "note": "Pages under /participants target research participants; all other pages target researchers"
    },
    "opentrain.ai": {
        "worker_path": "/become-freelancer/",
        "note": "Pages under /become-freelancer/ target workers; all other pages target clients"
    },
}

# Colours
C_HEADER      = "1F3864"   # dark navy
C_HEADER_FONT = "FFFFFF"
C_CLIENT      = "DDEEFF"   # light blue  — client
C_WORKER      = "EEFFDD"   # light green — worker
C_BOTH        = "FFF3CD"   # light amber — both / split
C_EMPTY       = "FFFFFF"
C_CODER_COL   = "FFF9C4"   # yellow — cells for coder to fill

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def header_cell(ws, row, col, value, width=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(bold=True, color=C_HEADER_FONT, name="Arial", size=10)
    c.fill      = PatternFill("solid", fgColor=C_HEADER)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = thin_border()
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c

def data_cell(ws, row, col, value="", fill=C_EMPTY, bold=False, wrap=False, align="left"):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", size=10, bold=bold)
    c.fill      = PatternFill("solid", fgColor=fill)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.border    = thin_border()
    return c

def row_fill(audience):
    if audience == "client":      return C_CLIENT
    if audience == "worker":      return C_WORKER
    if audience in ("both","split"): return C_BOTH
    return C_EMPTY


def load_domains(conn):
    """Load all platforms from DB, marking split domains."""
    rows = conn.execute("""
        SELECT DISTINCT
            pl.domain,
            pl.audience,
            pl.company_id,
            pl.platform_type,
            pl.hq_region,
            pl.name AS platform_name
        FROM platforms pl
        ORDER BY pl.audience, pl.domain
    """).fetchall()

    domains = []
    for r in rows:
        domain   = r["domain"]
        audience = r["audience"]
        is_split = any(domain.endswith(d) or domain == d
                       for d in URL_SPLITS.keys())
        domains.append({
            "domain":        domain,
            "audience":      "split" if is_split and audience == "both" else audience,
            "company_id":    r["company_id"] or "",
            "platform_type": r["platform_type"] or "",
            "hq_region":     r["hq_region"] or "",
            "platform_name": r["platform_name"] or "",
            "is_split":      is_split,
        })
    return domains


def build_instructions_sheet(wb):
    ws = wb.create_sheet("Instructions", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 120

    lines = [
        ("DarkSideofAI — Intercoder Reliability Coding Sheet", True, 14),
        ("", False, 11),
        ("TASK OVERVIEW", True, 11),
        ("This study analyses language used by AI data annotation platforms to address two different "
         "audiences: client-facing (B2B) platforms that sell services to businesses, and worker-facing "
         "(B2W) platforms that recruit human annotators. The audience classification is a foundational "
         "coding decision of the study.", False, 11),
        ("", False, 11),
        ("YOUR TASK", True, 11),
        ("Please review each website listed in the 'Domain Coding' sheet and independently assign an "
         "audience label. For the three websites in the 'URL Splits' sheet, please verify whether the "
         "URL-based splitting rule correctly separates client vs worker content.", False, 11),
        ("", False, 11),
        ("CODING CATEGORIES", True, 11),
        ("  client  — The website/page primarily addresses BUSINESSES or RESEARCHERS who purchase "
         "or procure AI data annotation services. Typical indicators: pricing for enterprise contracts, "
         "case studies for business clients, API documentation, SLA terms, ROI language.", False, 11),
        ("  worker  — The website/page primarily addresses HUMAN ANNOTATORS, crowd workers, or "
         "freelancers who perform data tasks for pay. Typical indicators: 'earn money', task "
         "instructions, payment terms, worker sign-up forms, community features.", False, 11),
        ("  both    — The website genuinely addresses both audiences equally with no clear "
         "separation. Use this sparingly — most sites have a dominant audience.", False, 11),
        ("  unclear — You cannot determine the audience from the website content. Please add a note.", False, 11),
        ("", False, 11),
        ("DECISION RULES", True, 11),
        ("1. Visit the live URL if possible. If not available, use the domain name and platform "
         "type column as context.", False, 11),
        ("2. Focus on the homepage and primary navigation — not individual deep-link pages.", False, 11),
        ("3. When in doubt between 'client' and 'both': if more than ~70% of visible content "
         "addresses one audience, code it as that audience.", False, 11),
        ("4. Paired platforms (same company_id): one should be client, one worker. If you code "
         "both as the same label, add a note explaining why.", False, 11),
        ("5. For URL splits (sheet 2): visit a sample of the URLs listed and confirm whether the "
         "split rule correctly separates the two audiences.", False, 11),
        ("", False, 11),
        ("HOW TO FILL IN THE SHEET", True, 11),
        ("Column 'Your Label' (yellow): enter one of: client / worker / both / unclear", False, 11),
        ("Column 'Notes' (yellow): optional — explain any doubts, edge cases, or disagreements.", False, 11),
        ("Leave other columns unchanged.", False, 11),
        ("", False, 11),
        ("CONTACT", True, 11),
        ("If you have questions about the coding scheme, contact: Yannick Kayser (yannick.kayser@gmail.com)", False, 11),
    ]

    for i, (text, bold, size) in enumerate(lines, start=1):
        c = ws.cell(row=i, column=1, value=text)
        c.font      = Font(name="Arial", size=size, bold=bold)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[i].height = 30 if bold else 45 if len(text) > 100 else 18


def build_domain_sheet(wb, domains):
    """
    Coder-facing sheet: NO study labels visible.
    Domain, platform name, platform type, company_id, hq_region are shown
    as neutral context. The coder fills in 'Your Label' and 'Notes' only.
    Study labels are stored in the hidden '_study_labels' sheet instead.
    """
    ws = wb.create_sheet("Domain Coding", 1)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    # 5 context columns + 2 coder-input columns — no study label column
    headers = [
        ("Domain / URL", 32),
        ("Platform Name", 22),
        ("Platform Type", 18),
        ("Company ID", 16),
        ("HQ Region", 12),
        ("Your Label\n(client / worker / both / unclear)", 22),
        ("Notes\n(optional)", 32),
    ]
    for col, (h, w) in enumerate(headers, 1):
        header_cell(ws, 1, col, h, width=w)
    ws.row_dimensions[1].height = 36

    for i, d in enumerate(domains, start=2):
        # All context cells use a uniform neutral background — no colour coding
        # that could hint at the expected answer
        data_cell(ws, i, 1, d["domain"],        fill="F7F7F7")
        data_cell(ws, i, 2, d["platform_name"], fill="F7F7F7")
        data_cell(ws, i, 3, d["platform_type"], fill="F7F7F7")
        data_cell(ws, i, 4, d["company_id"],    fill="F7F7F7")
        data_cell(ws, i, 5, d["hq_region"],     fill="F7F7F7", align="center")
        # Coder input columns — yellow
        data_cell(ws, i, 6, "", fill=C_CODER_COL)
        data_cell(ws, i, 7, "", fill=C_CODER_COL, wrap=True)


def build_hidden_labels_sheet(wb, domains):
    """
    Hidden sheet storing researcher's study labels for kappa calculation.
    The coder never sees this sheet — it is used by intercoder_reliability.py
    when comparing the two sets of labels after the coder returns the file.

    NOTE: openpyxl cannot truly password-protect a sheet, but the sheet is
    named with a leading underscore and kept at the end of the workbook so it
    is not the first thing a coder sees. For stronger protection, manually
    right-click the sheet tab in Excel → 'Hide' before sending the file.
    """
    ws = wb.create_sheet("_study_labels")  # underscore = internal / hidden convention
    ws.sheet_view.showGridLines = False

    header_cell(ws, 1, 1, "domain",        width=32)
    header_cell(ws, 1, 2, "study_audience", width=20)
    header_cell(ws, 1, 3, "is_split",       width=12)

    for i, d in enumerate(domains, start=2):
        ws.cell(row=i, column=1, value=d["domain"])
        ws.cell(row=i, column=2, value=d["audience"])
        ws.cell(row=i, column=3, value=str(d["is_split"]))

    # Move to last position so it doesn't show first
    wb.move_sheet("_study_labels", offset=len(wb.sheetnames))


def build_url_splits_sheet(wb, conn):
    """
    Coder-facing URL splits sheet.
    Shows domain + URL only — NO pre-assigned label, NO rationale.
    The coder assigns their own label independently.
    Researcher's URL-based labels are stored in '_study_labels_url' (hidden).
    """
    ws = wb.create_sheet("URL Coding", 2)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:E1")
    title = ws.cell(row=1, column=1,
        value="URL Coding — classify each page based on its content")
    title.font      = Font(bold=True, name="Arial", size=11, color=C_HEADER_FONT)
    title.fill      = PatternFill("solid", fgColor=C_HEADER)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = [
        ("Domain", 22),
        ("URL", 60),
        ("Your Label\n(client / worker / both / unclear)", 22),
        ("Notes\n(optional)", 35),
    ]
    for col, (h, w) in enumerate(headers, 1):
        header_cell(ws, 2, col, h, width=w)

    # Hidden sheet for researcher's URL labels
    ws_hidden = wb.create_sheet("_study_labels_url")
    header_cell(ws_hidden, 1, 1, "domain", width=22)
    header_cell(ws_hidden, 1, 2, "url",    width=60)
    header_cell(ws_hidden, 1, 3, "study_label", width=16)
    hidden_row = 2

    row = 3
    for domain, split in URL_SPLITS.items():
        pages = conn.execute("""
            SELECT p.url
            FROM   pages p
            JOIN   websites w ON w.id = p.website_id
            WHERE  w.domain LIKE ?
            ORDER  BY p.url
            LIMIT  40
        """, (f"%{domain}",)).fetchall()

        worker_path = split["worker_path"]

        if not pages:
            data_cell(ws, row, 1, domain,              fill="F7F7F7")
            data_cell(ws, row, 2, f"https://{domain}/", fill="F7F7F7", wrap=True)
            data_cell(ws, row, 3, "",                   fill=C_CODER_COL)
            data_cell(ws, row, 4, "",                   fill=C_CODER_COL, wrap=True)
            ws_hidden.cell(row=hidden_row, column=1, value=domain)
            ws_hidden.cell(row=hidden_row, column=2, value=f"https://{domain}/")
            ws_hidden.cell(row=hidden_row, column=3, value="client")
            hidden_row += 1
            row += 1
            continue

        for pg in pages:
            url        = pg[0]
            study_lbl  = "worker" if worker_path.lower() in url.lower() else "client"

            # Coder sees: domain + URL + empty input cells (neutral grey)
            data_cell(ws, row, 1, domain, fill="F7F7F7")
            data_cell(ws, row, 2, url,    fill="F7F7F7", wrap=True)
            data_cell(ws, row, 3, "",     fill=C_CODER_COL)
            data_cell(ws, row, 4, "",     fill=C_CODER_COL, wrap=True)

            # Hidden sheet stores researcher's label
            ws_hidden.cell(row=hidden_row, column=1, value=domain)
            ws_hidden.cell(row=hidden_row, column=2, value=url)
            ws_hidden.cell(row=hidden_row, column=3, value=study_lbl)
            hidden_row += 1
            row += 1

        # Spacer between domains
        for col in range(1, 5):
            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="EEEEEE")
        row += 1

    wb.move_sheet("_study_labels_url", offset=len(wb.sheetnames))


def export_csv(domains):
    csv_path = OUTPUT_DIR / "domain_list_for_codebook.csv"
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "domain", "platform_name", "platform_type",
            "company_id", "hq_region", "study_audience", "is_split"
        ])
        writer.writeheader()
        for d in domains:
            writer.writerow({
                "domain":         d["domain"],
                "platform_name":  d["platform_name"],
                "platform_type":  d["platform_type"],
                "company_id":     d["company_id"],
                "hq_region":      d["hq_region"],
                "study_audience": d["audience"],
                "is_split":       d["is_split"],
            })
    print(f"  CSV  → {csv_path}")


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row

    if not conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name='platforms'"
    ).fetchone():
        raise SystemExit("platforms table not found — run 01_prepare.py first.")

    print("Loading domains from database ...")
    domains = load_domains(conn)
    print(f"  {len(domains)} domains loaded")

    print("Building Excel workbook ...")
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    build_instructions_sheet(wb)
    build_domain_sheet(wb, domains)
    build_hidden_labels_sheet(wb, domains)
    build_url_splits_sheet(wb, conn)

    xlsx_path = OUTPUT_DIR / "intercoder_coding_sheet.xlsx"
    wb.save(xlsx_path)
    print(f"  XLSX → {xlsx_path}")

    export_csv(domains)
    conn.close()

    print("\nDone.")
    print("Next steps:")
    print("  1. Open intercoder_coding_sheet.xlsx and verify it looks correct")
    print("  2. Share it with the second coder alongside the codebook (.docx)")
    print("  3. After coding: run intercoder_reliability.py to compute Cohen's kappa")


if __name__ == "__main__":
    main()
