"""
intercoder_reliability.py
==========================
Calculates intercoder reliability (Cohen's kappa) after the second coder
returns the completed intercoder_coding_sheet.xlsx.

Usage
-----
    python3 src/intercoder_reliability.py

Reads:  output/intercoder/intercoder_coding_sheet.xlsx
Writes: output/intercoder/reliability_report.txt
"""

import sys
from pathlib import Path
from collections import Counter

try:
    import openpyxl
except ImportError:
    raise SystemExit("openpyxl not installed. Run: pip install openpyxl")

XLSX_PATH   = Path("output/intercoder/intercoder_coding_sheet_filledout.xlsx")
REPORT_PATH = Path("output/intercoder/reliability_report.txt")

VALID_LABELS = {"Client", "Worker", "Both", "Unclear"}

# Column indices — Domain Coding sheet (coder-facing, no study label)
COL_DOMAIN      = 1
COL_CODER_LABEL = 6   # coder's label
COL_NOTES       = 7

# Column indices — _study_labels hidden sheet
COL_HL_DOMAIN   = 1
COL_HL_LABEL    = 2


def cohen_kappa(rater1, rater2):
    """Compute Cohen's kappa for two lists of labels."""
    assert len(rater1) == len(rater2), "Rater lists must be the same length"
    n = len(rater1)
    if n == 0:
        return float("nan"), 0

    categories = sorted(set(rater1) | set(rater2))
    # Observed agreement
    p_o = sum(a == b for a, b in zip(rater1, rater2)) / n

    # Expected agreement
    p_e = sum(
        (rater1.count(c) / n) * (rater2.count(c) / n)
        for c in categories
    )

    if p_e == 1.0:
        return 1.0, n

    kappa = (p_o - p_e) / (1 - p_e)
    return round(kappa, 4), n


def interpret_kappa(k):
    if k < 0:       return "Poor (less than chance)"
    if k < 0.20:    return "Slight"
    if k < 0.40:    return "Fair"
    if k < 0.60:    return "Moderate"
    if k < 0.80:    return "Substantial  ✓ (meets thesis threshold)"
    return              "Almost perfect  ✓✓"


def main():
    if not XLSX_PATH.exists():
        raise SystemExit(f"File not found: {XLSX_PATH}\n"
                        "Run generate_intercoder_materials.py first and ensure "
                        "the second coder has returned their labels.")

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    # Load researcher's labels from hidden sheet
    if "_study_labels" not in wb.sheetnames:
        raise SystemExit(
            "Hidden sheet '_study_labels' not found.\n"
            "Re-run generate_intercoder_materials.py to rebuild the file."
        )
    study_label_map = {}
    for row in wb["_study_labels"].iter_rows(min_row=2, values_only=True):
        # Normalise to lowercase so lookup is case-insensitive
        domain = str(row[COL_HL_DOMAIN - 1] or "").strip().lower()
        label  = str(row[COL_HL_LABEL  - 1] or "").strip().lower()
        if domain:
            # Normalise split → both for kappa purposes
            study_label_map[domain] = "both" if label == "split" else label

    print(f"[DIAGNOSTIC] _study_labels sheet: {len(study_label_map)} domains loaded")
    if study_label_map:
        sample = list(study_label_map.items())[:5]
        print(f"[DIAGNOSTIC] First 5 hidden-sheet domains: {sample}")
    else:
        raise SystemExit(
            "ERROR: _study_labels sheet exists but is empty.\n"
            "Re-run generate_intercoder_materials.py to rebuild the file."
        )

    ws = wb["Domain Coding"]

    researcher_labels = []
    coder_labels      = []
    disagreements     = []
    missing_coder     = []
    skipped_no_study  = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        # Normalise domain to lowercase to match the hidden sheet key
        domain      = str(row[COL_DOMAIN - 1] or "").strip().lower()
        coder_label = str(row[COL_CODER_LABEL - 1] or "").strip().lower()
        note        = str(row[COL_NOTES - 1] or "").strip()

        if not domain:
            continue

        study_label = study_label_map.get(domain)
        if not study_label:
            skipped_no_study.append(domain)
            continue   # domain not in hidden sheet (legend row, header, etc.)

        if not coder_label or coder_label in ("none", "nan", ""):
            missing_coder.append(domain)
            continue

        # Validate coder label value
        valid_lower = {v.lower() for v in VALID_LABELS}
        if coder_label not in valid_lower:
            print(f"[WARNING] Unexpected coder label '{coder_label}' for '{domain}' — "
                  f"expected one of {sorted(valid_lower)}. Treating as missing.")
            missing_coder.append(domain)
            continue

        researcher_labels.append(study_label)
        coder_labels.append(coder_label)

        if study_label != coder_label:
            disagreements.append({
                "domain":     domain,
                "researcher": study_label,
                "coder":      coder_label,
                "note":       note,
            })

    kappa, n = cohen_kappa(researcher_labels, coder_labels)
    p_agree  = round(sum(a == b for a, b in zip(researcher_labels, coder_labels)) / max(n, 1) * 100, 1)

    # Diagnostic: show skipped rows if nothing matched — surface the root cause
    if n == 0:
        print("\n[DIAGNOSTIC] No domains matched between the two sheets.")
        print(f"  Domains in _study_labels : {len(study_label_map)}")
        print(f"  Rows in Domain Coding    : {ws.max_row - 1}")
        print(f"  Skipped (no match)       : {len(skipped_no_study)}")
        if skipped_no_study:
            print(f"  First 5 unmatched domain-coding rows: {skipped_no_study[:5]}")
        print("\nLikely cause: the domain strings in the two sheets differ.")
        print("Check for uppercase letters, 'www.' prefix, trailing spaces,")
        print("or extra columns shifting the domain column index.")
        raise SystemExit("Aborting — no data to compute kappa. See diagnostics above.")

    # ── Build report ──────────────────────────────────────────────────────────
    lines = [
        "DarkSideofAI — Intercoder Reliability Report",
        "=" * 55,
        "",
        f"Domains coded by both raters : {n}",
        f"Domains with missing coder label : {len(missing_coder)}",
        f"Domains skipped (not in hidden sheet) : {len(skipped_no_study)}",
        "",
        "OVERALL AGREEMENT",
        f"  Observed agreement (p_o) : {p_agree}%",
        f"  Cohen's kappa (κ)        : {kappa}",
        f"  Interpretation           : {interpret_kappa(kappa)}",
        "",
        "LABEL DISTRIBUTION",
        "  Researcher labels:",
    ]
    for label, count in sorted(Counter(researcher_labels).items()):
        lines.append(f"    {label:<10} {count:>4}  ({count/max(n,1)*100:.1f}%)")
    lines.append("  Coder labels:")
    for label, count in sorted(Counter(coder_labels).items()):
        lines.append(f"    {label:<10} {count:>4}  ({count/max(n,1)*100:.1f}%)")

    lines += [
        "",
        f"DISAGREEMENTS  ({len(disagreements)} cases)",
        "-" * 55,
    ]
    if disagreements:
        for d in disagreements:
            lines.append(f"  {d['domain']}")
            lines.append(f"    Researcher: {d['researcher']}  |  Coder: {d['coder']}")
            if d["note"]:
                lines.append(f"    Note: {d['note']}")
    else:
        lines.append("  None — perfect agreement.")

    if missing_coder:
        lines += ["", "DOMAINS WITHOUT CODER LABEL", "-" * 55]
        for d in missing_coder:
            lines.append(f"  {d}")

    lines += [
        "",
        "THESIS REPORTING TEMPLATE",
        "-" * 55,
        f"  'Intercoder reliability was assessed by having a second independent coder",
        f"  classify {n} platform domains using the same codebook. Cohen's kappa was",
        f"  κ = {kappa} ({interpret_kappa(kappa).split()[0].lower()} agreement), indicating",
        f"  {interpret_kappa(kappa).lower().replace('✓','').replace('✓✓','').strip()}.",
        f"  Disagreements ({len(disagreements)} cases, {round(len(disagreements)/max(n,1)*100,1)}%) were",
        f"  resolved through discussion between the two coders.'",
    ]

    report = "\n".join(lines)
    print(report)
    REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)
    REPORT_PATH.write_text(report, encoding="utf-8")
    print(f"\nReport saved → {REPORT_PATH}")


if __name__ == "__main__":
    main()
